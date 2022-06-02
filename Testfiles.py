import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import *
pd.options.mode.chained_assignment = None
import datetime
from tkinter.filedialog import askopenfilenames
import re
from tkinter import *
import os
import sys
w3 = Tk()
w3.title('CBN Validation Tool')
w3.geometry('300x450')
w3.minsize(300, 450)
w3.maxsize(300, 450)
greetings = Label(w3, text = "Please Choose CBN Duration")
greetings.pack(pady =10)
v = IntVar()
v.set(30)
scale = Scale(w3, variable=v, from_=1, to=30, orient=HORIZONTAL)
scale.pack(anchor=CENTER)
greetings2 = Label(w3, text = "",)
greetings3 = Label(w3, text = " Please Select Call Type")


print("                      Hi There !! Welcome")
def viewSelected():
    choice = var.get()
    if choice == 1:
        output_for_calls = "LTE Call, Redirect (normally) to test PSAP-UMTS Call, Redirect (normally) to test PSAP-GSM Call, Redirect (normally) to test PSAP"
        Calls_li = output_for_calls.split("-")
        return (Calls_li)


    elif choice == 2:
        output_for_calls = "LTE Call, Sent Network-Assisted GPS position to ALI:UMTS Call, Sent Network-Assisted GPS position to ALI:GSM Call, Sent Network-Assisted GPS position to ALI:LTE Call, OK:UMTS Call, OK:GSM Call, OK:LTE Call, Sent CELL location to ALI:UMTS Call, Sent CELL location to ALI:GSM Call, Sent CELL location to ALI"
        Calls_li = output_for_calls.split(":")
        return (Calls_li)
    elif choice == 3:
        output_for_calls = "LTE Call, Sent Network-Assisted GPS position to ALI:UMTS Call, Sent Network-Assisted GPS position to ALI:GSM Call, Sent Network-Assisted GPS position to ALI:LTE Call, OK:UMTS Call, OK:GSM Call, OK:LTE Call, Sent CELL location to ALI:UMTS Call, Sent CELL location to ALI:GSM Call, Sent CELL location to ALI:LTE Call, Redirect (normally) to test PSAP:UMTS Call, Redirect (normally) to test PSAP:GSM Call, Redirect (normally) to test PSAP"
        Calls_li = output_for_calls.split(":")
        return(Calls_li)

greetings2.pack()
greetings3.pack(pady=10)

var = IntVar()
Radiobutton(w3,text="1. PSAP Calls            ", variable=var, value=1, command=viewSelected).pack(pady=5)
Radiobutton(w3,text="2. Live Calls              ", variable=var, value=2, command=viewSelected).pack(pady=5)
Radiobutton(w3,text="3. PSAP + Live Calls", variable=var, value=3, command=viewSelected).pack(pady=5)
greet = Label(w3, text="", ).pack()

# Button(w3,text="Submit",bg= "light grey",command = w3.destroy).pack(pady =25)
Label(w3, text="V4.1 ---- By Sumit Kamboj", font = "Mistral 12 ",).pack(fill=X, side = "bottom")
Label(w3,text = "Amdocs@2021", font = "lucia 6").pack(side="bottom")
import tkinter as tk
def printtext():
    global e1
    string2 = e1.get()
    return string2
tk.Label(w3,text="Enter SiteID: ").pack()
e1 = tk.Entry(w3)
e1.pack()
# e = Entry(w3)
# name_label = tk.Label(w3, text = 'Enter SiteID: ', font=('calibre',10, 'bold'))
# e.pack()
e1.focus_set()

def import_file(func):
    def inner():
        exlpath = askopenfilenames(title = "Open 'csv' file")
        import_out = pd.concat([pd.read_csv(f) for f in exlpath])
        print(f"CSV file loaded successfully")
        return func(import_out)
    return inner
@import_file
def main_program(filename):

    call_type = viewSelected()

    if call_type == None:
        raise Exception("You have not seleted any call type please select call type from options and try again")
    file_new = filename[(filename.duration > v.get() ) & (filename.endStatusExplan.isin(viewSelected()))]
    file_new["SiteName4"] = file_new.loc[:,("siteId")].str[:8].copy(deep=True)
    print(f"You Selected Call Duration >= {v.get()} sec")
    var2 = printtext().upper()


    if len(call_type) == 3:
        print(f'You Selected Call type option 1. PSAP Calls')
    if len(call_type) == 9:
        print(f'You selected Call type option 2. Live Calls')
    if len(call_type) == 12:
        print(f'You selected call type option 3. PSAP + Live Calls')

    print(f'You entered siteID: {var2}')




    if (len(var2) == 0):
        raise Exception("sitedID can not be blank, Enter valid siteID and import file again...")
    if (len(var2) != 8):
        raise Exception("Enter valid siteID and import file again...")



    file_new2 = file_new[(file_new.SiteName4 == var2)]
    if file_new2.empty:
        raise Exception(f"Input csv files do not have any call records for Entered siteID {var2}..please verify and try again")

    filename["SiteName5"] = filename.loc[:,("siteId")].str[:8].copy(deep=True)
    filename['cgi2'] = filename['cgi'].map(lambda x: re.sub("[-]", '', x))
    filename["CGI check"] = filename["siteId"] + "&" + filename["cgi2"]

    file_new8 = filename[(filename.SiteName5 == var2)]



    file_new2["SiteName"] = file_new2.loc[:,("siteId")].str[-5:].copy(deep=True)
    file_new2["SiteName2"] = file_new2.loc[:,("siteId")].str[-5:].copy(deep=True)
    file_new2['SiteName'] = file_new2['SiteName'].str.replace('11LFA', 'AWS3-A').str.replace('21LFA', 'AWS3-B')	.str.replace('31LFA', 'AWS3-C')	.str.replace('41LFA', 'AWS3-D')	.str.replace('51LFA', 'AWS3-E')	.str.replace('61LFA', 'AWS3-Z').str.replace('11LEA', 'L600-A')	.str.replace('21LEA', 'L600-B')	.str.replace('31LEA', 'L600-C')	.str.replace('41LEA', 'L600-D')	.str.replace('51LEA', 'L600-E')	.str.replace('61LEA', 'L600-Z').str.replace('11LDA', 'L700-A')	.str.replace('21LDA', 'L700-B')	.str.replace('31LDA', 'L700-C')	.str.replace('41LDA', 'L700-D')	.str.replace('51LDA', 'L700-E')	.str.replace('61LDA', 'L700-Z').str.replace('11LPA', 'L1900_1C-A')	.str.replace('21LPA', 'L1900_1C-B')	.str.replace('31LPA', 'L1900_1C-C')	.str.replace('41LPA', 'L1900_1C-D')	.str.replace('51LPA', 'L1900_1C-E')	.str.replace('61LPA', 'L1900_1C-Z').str.replace('12LPA', 'L1900_2C-A')	.str.replace('22LPA', 'L1900_2C-B')	.str.replace('32LPA', 'L1900_2C-C')	.str.replace('42LPA', 'L1900_2C-D')	.str.replace('52LPA', 'L1900_2C-E')	.str.replace('62LPA', 'L1900_2C-Z').str.replace('11LAA', 'L2100_1C-A')	.str.replace('21LAA', 'L2100_1C-B')	.str.replace('31LAA', 'L2100_1C-C')	.str.replace('41LAA', 'L2100_1C-D')	.str.replace('51LAA', 'L2100_1C-E')	.str.replace('61LAA', 'L2100_1C-Z').str.replace('12LAA', 'L2100_2C-A')	.str.replace('22LAA', 'L2100_2C-B')	.str.replace('32LAA', 'L2100_2C-C')	.str.replace('42LAA', 'L2100_2C-D')	.str.replace('52LAA', 'L2100_2C-E')	.str.replace('62LAA', 'L2100_2C-Z').str.replace('11LKA', 'L2500_1C-A')	.str.replace('21LKA', 'L2500_1C-B')	.str.replace('31LKA', 'L2500_1C-C')	.str.replace('41LKA', 'L2500_1C-D')	.str.replace('51LKA', 'L2500_1C-E')	.str.replace('61LKA', 'L2500_1C-Z').str.replace('12LKA', 'L2500_2C-A')	.str.replace('22LKA', 'L2500_2C-B')	.str.replace('32LKA', 'L2500_2C-C')	.str.replace('42LKA', 'L2500_2C-D')	.str.replace('52LKA', 'L2500_2C-E')	.str.replace('62LKA', 'L2500_2C-Z').str.replace('13LKA', 'L2500_3C-A')	.str.replace('23LKA', 'L2500_3C-B')	.str.replace('33LKA', 'L2500_3C-C')	.str.replace('43LKA', 'L2500_3C-D')	.str.replace('53LKA', 'L2500_3C-E')	.str.replace('63LKA', 'L2500_3C-Z').str.replace('11UPA', 'U1900_1C-A')	.str.replace('21UPA', 'U1900_1C-B')	.str.replace('31UPA', 'U1900_1C-C')	.str.replace('41UPA', 'U1900_1C-D')	.str.replace('51UPA', 'U1900_1C-E')	.str.replace('61UPA', 'U1900_1C-Z').str.replace('12UPA', 'U1900_2C-A')	.str.replace('22UPA', 'U1900_2C-B')	.str.replace('32UPA', 'U1900_2C-C')	.str.replace('42UPA', 'U1900_2C-D')	.str.replace('52UPA', 'U1900_2C-E')	.str.replace('62UPA', 'U1900_2C-Z').str.replace('11UAA', 'U2100_1C-A')	.str.replace('21UAA', 'U2100_1C-B')	.str.replace('31UAA', 'U2100_1C-C')	.str.replace('41UAA', 'U2100_1C-D')	.str.replace('51UAA', 'U2100_1C-E')	.str.replace('61UAA', 'U2100_1C-Z').str.replace('12UAA', 'U2100_2C-A')	.str.replace('22UAA', 'U2100_2C-B')	.str.replace('32UAA', 'U2100_2C-C')	.str.replace('42UAA', 'U2100_2C-D')	.str.replace('52UAA', 'U2100_2C-E')	.str.replace('62UAA', 'U2100_2C-Z').str.replace('A0GPA', 'GSM-A')	.str.replace('B0GPA', 'GSM-B')	.str.replace('C0GPA', 'GSM-C')	.str.replace('D0GPA', 'GSM-D')	.str.replace('E0GPA', 'GSM-E')	.str.replace('F0GPA', 'GSM-Z')

    file_new2['SiteName2'] = file_new2['SiteName2'].str.replace('11LFA', 'AWS3-A').str.replace('21LFA', 'AWS3-B')	.str.replace('31LFA', 'AWS3-G')	.str.replace('41LFA', 'AWS3-D')	.str.replace('51LFA', 'AWS3-E')	.str.replace('61LFA', 'AWS3-Z').str.replace('11LEA', 'L600-A')	.str.replace('21LEA', 'L600-B')	.str.replace('31LEA', 'L600-G')	.str.replace('41LEA', 'L600-D')	.str.replace('51LEA', 'L600-E')	.str.replace('61LEA', 'L600-Z').str.replace('11LDA', 'L700-A')	.str.replace('21LDA', 'L700-B')	.str.replace('31LDA', 'L700-G')	.str.replace('41LDA', 'L700-D')	.str.replace('51LDA', 'L700-E')	.str.replace('61LDA', 'L700-Z').str.replace('11LPA', 'L1900_1C-A')	.str.replace('21LPA', 'L1900_1C-B')	.str.replace('31LPA', 'L1900_1C-G')	.str.replace('41LPA', 'L1900_1C-D')	.str.replace('51LPA', 'L1900_1C-E')	.str.replace('61LPA', 'L1900_1C-Z').str.replace('12LPA', 'L1900_2C-A')	.str.replace('22LPA', 'L1900_2C-B')	.str.replace('32LPA', 'L1900_2C-G')	.str.replace('42LPA', 'L1900_2C-D')	.str.replace('52LPA', 'L1900_2C-E')	.str.replace('62LPA', 'L1900_2C-Z').str.replace('11LAA', 'L2100_1C-A')	.str.replace('21LAA', 'L2100_1C-B')	.str.replace('31LAA', 'L2100_1C-G')	.str.replace('41LAA', 'L2100_1C-D')	.str.replace('51LAA', 'L2100_1C-E')	.str.replace('61LAA', 'L2100_1C-Z').str.replace('12LAA', 'L2100_2C-A')	.str.replace('22LAA', 'L2100_2C-B')	.str.replace('32LAA', 'L2100_2C-G')	.str.replace('42LAA', 'L2100_2C-D')	.str.replace('52LAA', 'L2100_2C-E')	.str.replace('62LAA', 'L2100_2C-Z').str.replace('11LKA', 'L2500_1C-A')	.str.replace('21LKA', 'L2500_1C-B')	.str.replace('31LKA', 'L2500_1C-G')	.str.replace('41LKA', 'L2500_1C-D')	.str.replace('51LKA', 'L2500_1C-E')	.str.replace('61LKA', 'L2500_1C-Z').str.replace('12LKA', 'L2500_2C-A')	.str.replace('22LKA', 'L2500_2C-B')	.str.replace('32LKA', 'L2500_2C-G')	.str.replace('42LKA', 'L2500_2C-D')	.str.replace('52LKA', 'L2500_2C-E')	.str.replace('62LKA', 'L2500_2C-Z').str.replace('13LKA', 'L2500_3C-A')	.str.replace('23LKA', 'L2500_3C-B')	.str.replace('33LKA', 'L2500_3C-G')	.str.replace('43LKA', 'L2500_3C-D')	.str.replace('53LKA', 'L2500_3C-E')	.str.replace('63LKA', 'L2500_3C-Z').str.replace('11UPA', 'U1900_1C-A')	.str.replace('21UPA', 'U1900_1C-B')	.str.replace('31UPA', 'U1900_1C-G')	.str.replace('41UPA', 'U1900_1C-D')	.str.replace('51UPA', 'U1900_1C-E')	.str.replace('61UPA', 'U1900_1C-Z').str.replace('12UPA', 'U1900_2C-A')	.str.replace('22UPA', 'U1900_2C-B')	.str.replace('32UPA', 'U1900_2C-G')	.str.replace('42UPA', 'U1900_2C-D')	.str.replace('52UPA', 'U1900_2C-E')	.str.replace('62UPA', 'U1900_2C-Z').str.replace('11UAA', 'U2100_1C-A')	.str.replace('21UAA', 'U2100_1C-B')	.str.replace('31UAA', 'U2100_1C-G')	.str.replace('41UAA', 'U2100_1C-D')	.str.replace('51UAA', 'U2100_1C-E')	.str.replace('61UAA', 'U2100_1C-Z').str.replace('12UAA', 'U2100_2C-A')	.str.replace('22UAA', 'U2100_2C-B')	.str.replace('32UAA', 'U2100_2C-G')	.str.replace('42UAA', 'U2100_2C-D')	.str.replace('52UAA', 'U2100_2C-E')	.str.replace('62UAA', 'U2100_2C-Z').str.replace('A0GPA', 'GSM-A')	.str.replace('B0GPA', 'GSM-B')	.str.replace('C0GPA', 'GSM-G')	.str.replace('D0GPA', 'GSM-D')	.str.replace('E0GPA', 'GSM-E')	.str.replace('F0GPA', 'GSM-Z')




    file_new2["status"]= file_new2.loc[:,("SiteName")].str[:].copy(deep=True)
    file_new2["Result"]= file_new2.loc[:,("SiteName2")].str[:].copy(deep=True)


    file_new3 = file_new2.drop_duplicates(subset=['status'])
    file_new4 = file_new3.sort_values(by=['status'])
    file_new5 = file_new2.sort_values(by=['status'])
    file_new9 = file_new8.drop_duplicates(subset='CGI check')
    file_new9["Cell with Duplicate CGI"] = file_new9.loc[:,("CGI check")].str[:14].copy(deep=True)
    file_new10 = file_new9[file_new9.duplicated(['Cell with Duplicate CGI'])]
    list_of_site = [' '.join([row for row in file_new4['Result']])]
    a2 = ' '.join(map(str, list_of_site))
    site_list = a2.split(" ")
    subs1 = "AWS3"
    subs2 = "L600"
    subs3 = "L700"
    subs2 = "L600"
    subs3 = "L700"
    subs4 = "L1900_1C"
    subs5 = "L1900_2C"
    subs6 = "L2100_1C"
    subs7 = "L2100_2C"
    subs8 = "L2500_1C"
    subs9 = "L2500_2C"
    subs10 = "L2500_3C"
    subs11 = "U1900_1C"
    subs12 = "U1900_2C"
    subs13 = "U2100_1C"
    subs14 = "U2100_2C"
    subs15 = "GSM"

    output1 =  [i for i in site_list if subs1 in i]
    output1= ' '.join(map(str, output1))
    output1 = output1.replace("AWS3-", "", 6).replace(" ", "")
    output_AWS3 = "AWS3-" + output1

    output2 =  [i for i in site_list if subs2 in i]
    output2= ' '.join(map(str, output2))
    output2 = output2.replace("L600-", "", 6).replace(" ", "")
    output_L600 = "L600-" + output2

    output3 =  [i for i in site_list if subs3 in i]
    output3= ' '.join(map(str, output3))
    output3 = output3.replace("L700-", "", 6).replace(" ", "")
    output_L700 = "L700-" + output3

    output4 =  [i for i in site_list if subs4 in i]
    output4= ' '.join(map(str, output4))
    output4 = output4.replace("L1900_1C-", "", 6).replace(" ", "")
    output_L1900_1C = "L1900_1C-" + output4

    output5 =  [i for i in site_list if subs5 in i]
    output5= ' '.join(map(str, output5))
    output5 = output5.replace("L1900_2C-", "", 6).replace(" ", "")
    output_L1900_2C = "L1900_2C-" + output5

    output6 =  [i for i in site_list if subs6 in i]
    output6= ' '.join(map(str, output6))
    output6 = output6.replace("L2100_1C-", "", 6).replace(" ", "")
    output_L2100_1C = "L2100_1C-" + output6

    output7 =  [i for i in site_list if subs7 in i]
    output7= ' '.join(map(str, output7))
    output7 = output7.replace("L2100_2C-", "", 6).replace(" ", "")
    output_L2100_2C = "L2100_2C-" + output7

    output8 =  [i for i in site_list if subs8 in i]
    output8= ' '.join(map(str, output8))
    output8 = output8.replace("L2500_1C-", "", 6).replace(" ", "")
    output_L2500_1C = "L2500_1C-" + output8

    output9 =  [i for i in site_list if subs9 in i]
    output9 = ' '.join(map(str, output9))
    output9 = output9.replace("L2500_2C-", "", 6).replace(" ", "")
    output_L2500_2C = "L2500_2C-" + output9

    output10 =  [i for i in site_list if subs10 in i]
    output10 = ' '.join(map(str, output10))
    output10 = output10.replace("L2500_3C-", "", 6).replace(" ", "")
    output_L2500_3C = "L2500_3C-" + output10

    output11 =  [i for i in site_list if subs11 in i]
    output11= ' '.join(map(str, output11))
    output11 = output11.replace("U1900_1C-", "", 6).replace(" ", "")
    output_U1900_1C = "U1900_1C-" + output11

    output12 = [i for i in site_list if subs12 in i]
    output12 = ' '.join(map(str, output12))
    output12 = output12.replace("U1900_2C-", "", 6).replace(" ", "")
    output_U1900_2C = "U1900_2C-" + output12

    output13 = [i for i in site_list if subs13 in i]
    output13 = ' '.join(map(str, output13))
    output13 = output13.replace("U2100_1C-", "", 6).replace(" ", "")
    output_U2100_1C = "U2100_1C-" + output13

    output14 = [i for i in site_list if subs14 in i]
    output14 = ' '.join(map(str, output14))
    output14 = output14.replace("U2100_2C-", "", 6).replace(" ", "")
    output_U2100_2C = "U2100_2C-" + output14
    output15 = [i for i in site_list if subs15 in i]
    output15 = ' '.join(map(str, output15))
    output15 = output15.replace("GSM-", "", 6).replace(" ", "")
    output_GSM = "GSM-" + output15
    title = file_new2["siteId"].iloc[0]


    try:
      os.mkdir('Output Files')
    except Exception:
        pass



    name = f'{title[:8]}_Result_{datetime.datetime.now().strftime("%H%M_%d%m%Y")}.xlsx'
    name2 = 'site_id_' + title[:8] + f'_{datetime.datetime.now().strftime("%M%S")}.csv'

    writer = pd.ExcelWriter(os.path.join('Output Files', name), engine='xlsxwriter')


    file_new8[['callStartTimeGMT',	"cbn",	"psapId","psapName","phase","state","esrk",	"siteId","cgi","locationDesc","initialLatLong","updatedLatLongAlt","metersBetweenPositions","positionMethod","positionSource","numAliQuery","mscName","imeiModel",	"imsi",	"duration",	"endStatusExplan",	"psapSelectType"]].to_csv(os.path.join('Output Files',name2),index=False)
    file_new5[['siteId', 'cbn', 'duration', 'endStatusExplan', 'Result']].to_excel(writer,sheet_name='CBN Result',index= False)
    file_new10[['Cell with Duplicate CGI']].to_excel(writer,sheet_name='CGI Verification',index= False)

    writer.save()

    wb = load_workbook(os.path.join('Output Files', name))

    wb.save(os.path.join('Output Files', name))
    wb.create_sheet('CBN Table')
    wb.save(os.path.join('Output Files', name))

    ws = wb["CBN Table"]
    ws["A1"] = "RIOT/CBN Check"
    ws["B1"] = "L600"
    ws["C1"] = "L700"
    ws["D1"] = "L19 1st Carrier"
    ws["E1"] = "L19 2nd Carrier"
    ws["F1"] = "L21_1st Carrier"
    ws["G1"] = "L21 2nd Carrier"
    ws["H1"] = "L25 1st carrier"
    ws["I1"] = "L25 2nd carrier"
    ws["J1"] = "L25 3rd carrier"
    ws["K1"] = "AWS3"
    ws["L1"] = "N600"
    ws["M1"] = "N25"
    ws["N1"] = "U19_1st carrier"
    ws["O1"] = "U19_2nd carrier"
    ws["P1"] = "U21_1st carrier"
    ws["Q1"] = "U21_2nd carrier"
    ws["R1"] = "GSM"
    ws["A2"] = "RIOT green"
    ws["A3"] = "RIOT red"
    ws["A4"] = "CBN Status"

    thin = Side(border_style="thick", color="000000")  # border style, color
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws['A3'].border = border

    for row in ws['A1:R4']:
        for cell in row:
            cell.border = border

    greenFill = PatternFill(start_color='92D059',
                            end_color='92D059',
                            fill_type='solid')
    ws['A2'].fill = greenFill

    redFill = PatternFill(start_color='FFFF0000',
                          end_color='FFFF0000',
                          fill_type='solid')
    ws['A3'].fill = redFill

    gray = PatternFill(start_color='BFBFBF',
                       end_color='BFBFBF',
                       fill_type='solid')
    ws['A4'].fill = gray

    for row in ws['A1:R1']:
        for cell in row:
            cell.fill = gray

    wb.save(os.path.join('Output Files', name))
    ws['A4'].font = Font(bold=True)
    ws['A2'].font = Font(bold=True)
    ws['A3'].font = Font(bold=True)

    for row in ws['A1:Q1']:
        for cell in row:
            cell.font = Font(bold=True)

    ws.merge_cells('A6:B6')

    ws["A6"] = 'CBN is Pass for:'

    if len(output_L600[5:]) > 0:
        ws["B4"] = output_L600[5:] + " Pass"
    if len(output_L700[5:]) > 0:
        ws["C4"] = output_L700[5:] + " Pass"
    if len(output_L1900_1C[9:]) > 0:
        ws["D4"] = output_L1900_1C[9:] + " Pass"
    if len(output_L1900_2C[9:]) > 0:
        ws["E4"] = output_L1900_2C[9:] + " Pass"
    if len(output_L2100_1C[9:]) > 0:
        ws["F4"] = output_L2100_1C[9:] + " Pass"
    if len(output_L2100_2C[9:]) > 0:
        ws["G4"] = output_L2100_2C[9:] + " Pass"
    if len(output_L2500_1C[9:]) > 0:
        ws["H4"] = output_L2500_1C[9:] + " Pass"
    if len(output_L2500_2C[9:]) > 0:
        ws["I4"] = output_L2500_2C[9:] + " Pass"
    if len(output_L2500_3C[9:]) > 0:
        ws["J4"] = output_L2500_3C[9:] + " Pass"
    if len(output_AWS3[5:]) > 0:
        ws["K4"] = output_AWS3[5:] + " Pass"
    if len(output_U1900_1C[9:]) > 0:
        ws["N4"] = output_U1900_1C[9:] + " Pass"
    if len(output_U1900_2C[9:]) > 0:
        ws["O4"] = output_U1900_2C[9:] + " Pass"
    if len(output_U2100_1C[9:]) > 0:
        ws["P4"] = output_U2100_1C[9:] + " Pass"
    if len(output_U2100_2C[9:]) > 0:
        ws["Q4"] = output_U2100_2C[9:] + " Pass"
    if len(output_GSM[4:]) > 0:
        ws["R4"] = output_GSM[4:] + " Pass"

    li_L600 = list(output_L600.split("-"))
    ws["B7"] = li_L600[1]
    if len(li_L600[1]) > 0:
         ws["A7"] = "L600"

    li_L700 = list(output_L700.split("-"))
    ws["B8"] = li_L700[1]
    if len(li_L700[1]) > 0:
            ws["A8"] = "L700"

    li_L1900_1C = list(output_L1900_1C.split("-"))
    ws["B9"] = li_L1900_1C[1]
    if len(li_L1900_1C[1]) > 0:
            ws["A9"] = "L1900_1C"

    li_L1900_2C = list(output_L1900_2C.split("-"))
    ws["B10"] = li_L1900_2C[1]
    if len(li_L1900_2C[1]) > 0:
            ws["A10"] = "L1900_2C"

    li_L2100_1C = list(output_L2100_1C.split("-"))
    ws["B11"] = li_L2100_1C[1]
    if len(li_L2100_1C[1]) > 0:
            ws["A11"] = "L2100_1C"

    li_L2100_2C = list(output_L2100_2C.split("-"))
    ws["B12"] = li_L2100_2C[1]
    if len(li_L2100_2C[1]) > 0:
            ws["A12"] = "L2100_2C"

    li_L2500_1C = list(output_L2500_1C.split("-"))
    ws["B13"] = li_L2500_1C[1]
    if len(li_L2500_1C[1]) > 0:
            ws["A13"] = "L2500_1C"

    li_L2500_2C = list(output_L2500_2C.split("-"))
    ws["B14"] = li_L2500_2C[1]
    if len(li_L2500_2C[1]) > 0:
            ws["A14"] = "L2500_2C"

    li_L2500_3C = list(output_L2500_3C.split("-"))
    ws["B15"] = li_L2500_3C[1]
    if len(li_L2500_3C[1]) > 0:
            ws["A15"] = "L2500_3C"

    li_AWS3 = list(output_AWS3.split("-"))
    ws["B16"] = li_AWS3[1]
    if len(li_AWS3[1]) > 0:
            ws["A16"] = "AWS3"

    li_U1900_1C = list(output_U1900_1C.split("-"))
    ws["B17"] = li_U1900_1C[1]
    if len(li_U1900_1C[1]) > 0:
            ws["A17"] = "U1900_1C"

    li_U1900_2C = list(output_U1900_2C.split("-"))
    ws["B18"] = li_U1900_2C[1]
    if len(li_U1900_2C[1]) > 0:
            ws["A18"] = "U1900_2C"

    li_U2100_1C = list(output_U2100_1C.split("-"))
    ws["B19"] = li_U2100_1C[1]
    if len(li_U2100_1C[1]) > 0:
            ws["A19"] = "U2100_1C"

    li_U2100_2C = list(output_U2100_2C.split("-"))
    ws["B20"] = li_U2100_2C[1]
    if len(li_U2100_2C[1]) > 0:
            ws["A20"] = "U2100_2C"

    li_GSM = list(output_GSM.split("-"))
    ws["B21"] = li_GSM[1]
    if len(li_GSM[1]) > 0:
            ws["A21"] = "GSM"

    ws["A5"] = " "

    wb.save(os.path.join('Output Files', name))

    index_row = []

    # loop each row in column A
    for i in range(1, ws.max_row):
        # define emptiness of cell
        if ws.cell(i, 1).value is None:
            # collect indexes of rows
            index_row.append(i)

    # loop each index value
    for row_del in range(len(index_row)):
        ws.delete_rows(idx=index_row[row_del], amount=1)
        # exclude offset of rows through each iteration
        index_row = list(map(lambda k: k - 1, index_row))


    ws.delete_rows
    wb.save(os.path.join('Output Files', name))
    print(f'Output file generated successfully for site -----> {title[:8]}')
    print(f'\n')

btn4 = Label(w3, text="", ).pack()
btn5 = Label(w3, text="", ).pack()
btn = Button(w3, text='Import csv files..' ,command = main_program,bg='grey')
btn.pack(padx=50)
w3.mainloop()







